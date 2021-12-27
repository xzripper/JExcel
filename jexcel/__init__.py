from typing import Union
from openpyxl import Workbook
from json import load
from os import remove


class JExcel:
    JEXCEL_VERSION = 1.0

    def __init__(self, json: Union[None, str]=None, out: Union[None, str]=None) -> None:
        """Create excel file with keys and value from json."""
        self.json = json
        self.out = out

    def setjson(self, json: str) -> None:
        """Update json file."""
        self.json = json

    def setout(self, out: str) -> None:
        """Update out file."""
        self.out = out

    def getexcel(self) -> Union[None, str]:
        """Get excel file with extension."""
        if self.out is None:
            return f'{self.json}.xlsx'

        else:
            return f'{self.out}.xlsx'

    def getjson(self) -> Union[None, str]:
        """Get json file with extension."""
        if self.json is None:
            raise TypeError('json is not specified')

        else:
            return f'{self.json}.json'

    def create(self, msg: Union[bool, str]=False) -> None:
        """Create excel with data."""
        excelfile = self.getexcel()
        jsonfile = self.getjson()

        content = load(open(jsonfile, 'r'))

        excel = Workbook()
        active = excel.active

        cell = 2

        if type(msg) is bool:
            if msg:
                active.cell(1, 1, f'{self.getjson()} settings.')

            elif not msg:
                cell -= 1

        elif type(msg) is str:
            active.cell(1, 1, msg)

        else:
            raise TypeError(f'bool | str needed, {type(msg).__name__} found')

        for key, num in zip(content, range(cell, len(content.keys()) + 2)):
            active.cell(num, 1, str(key))

            if type(content[key]) is list:
                beautiful = (str(', '.join(list(map(lambda element: str(element), content[key])))))

                active.cell(num, 2, beautiful)

            elif type(content[key]) is dict:
                _nwl = '\n'
                beautiful = ''

                for _key in content[key]:
                    # TODO: Add nested dicts support.
                    if type(content[key][_key]) is dict:
                        raise OverflowError('too many dicts')

                    else:
                        if list(content[key].keys())[-1] == _key:
                            formatted = f'{_key} = {str(content[key][_key]).replace(_nwl, "")}'

                        else:
                            formatted = f'{_key} = {str(content[key][_key]).replace(_nwl, "")}, '

                        beautiful += formatted 

                    active.cell(num, 2, beautiful)

            else:
                active.cell(num, 2, str(content[key]))

        excel.save(excelfile)

    def remove(self) -> None:
        """Remove excel file."""
        remove(self.getexcel())

def jev() -> float:
    """Get version of library."""
    return JExcel.JEXCEL_VERSION
